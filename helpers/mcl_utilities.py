import os
import math
import glob
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import Workbook

from openpyxl.utils.dataframe import dataframe_to_rows

from sklearn.metrics import (
    confusion_matrix, accuracy_score, precision_score, recall_score, f1_score,
    roc_curve, auc, precision_recall_curve, average_precision_score
)


def save_metrics_to_excel(output_path, gam_df, method_df, class_df=None):
    """Write DataFrames to an Excel workbook."""
    wb = Workbook()

    # API attentions
    ws1 = wb.active
    ws1.title = "api_attentions"
    for row in dataframe_to_rows(gam_df, index=False, header=True):
        ws1.append(row)
    ws1.auto_filter.ref = ws1.dimensions
    ws1.column_dimensions['A'].width = 70
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20

    # Method localization
    ws2 = wb.create_sheet(title="method_loc")
    for row in dataframe_to_rows(method_df, index=False, header=True):
        ws2.append(row)
    ws2.auto_filter.ref = ws2.dimensions
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15
    ws2.column_dimensions['G'].width = 15

    # Class localization (optional)
    if class_df is not None:
        ws3 = wb.create_sheet(title="class_loc")
        for row in dataframe_to_rows(class_df, index=False, header=True):
            ws3.append(row)
        ws3.auto_filter.ref = ws3.dimensions
        ws3.column_dimensions['A'].width = 70
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 20
        ws3.column_dimensions['D'].width = 20

    # Save workbook
    wb.save(output_path)


def kaspersky_to_folder_name(apk_name, kaspersky_value):
    if apk_name.startswith("mw_androzoo_"):
        return "ood"
    else:
        main_part = kaspersky_value.split(':')[-1]
        parts = main_part.split('.')
        folder_name = '_'.join(parts[-2:]).lower()
        return folder_name


def find_common_filenames(path_1, path_2, path_3, path_4, path_5):
    # Get a list of filenames in each folder
    mcl_baselines = os.listdir(path_1)
    gam_files = os.listdir(path_2)
    gat_files = os.listdir(path_3)
    custom_files = os.listdir(path_4)
    custom_methods_and_called_apis_files = os.listdir(path_5)

    # Extract filenames without extensions
    mcl_baselines = [os.path.splitext(filename)[0] for filename in mcl_baselines]
    gam_filenames = [os.path.splitext(filename)[0] for filename in gam_files]
    gat_filenames = [os.path.splitext(filename)[0] for filename in gat_files]
    custom_filenames = [os.path.splitext(filename)[0] for filename in custom_files]
    custom_methods_and_called_apis_filenames = [os.path.splitext(filename)[0] for filename in
                                                custom_methods_and_called_apis_files]

    # Find common filenames using set intersection
    common_file_names = set(mcl_baselines) & set(gam_filenames) & set(gat_filenames) & \
                        set(custom_filenames) & set(custom_methods_and_called_apis_filenames)
    print(f"Total number of mcl_baselines: {len(mcl_baselines)}")
    print(f"Total number of gam_filenames: {len(gam_filenames)}")
    print(f"Total number of gat_filenames: {len(gat_filenames)}")
    print(f"Total number of custom_filenames: {len(custom_filenames)}")
    print(f"Total number of custom_methods_and_called_apis_filenames: {len(custom_methods_and_called_apis_filenames)}")
    print(f"Total number of files: {len(common_file_names)}")
    return common_file_names


def process_apk_methods_and_classes(
        apk_filename,
        gam_attention_path, gat_attention_path,
        custom_methods_and_called_apis_path,
        custom_methods_path, class_mcl_baseline_path, method_mcl_baseline_path,
        excluded_classes, number_of_top_apis
):
    # --- Load ground truths ---
    real_malicious_method_set = get_real_malicious_set(method_mcl_baseline_path, apk_filename)
    real_malicious_class_set = get_real_malicious_set(class_mcl_baseline_path, apk_filename)

    # --- Load CSV (methods + called APIs) ---
    csv_path = os.path.join(custom_methods_and_called_apis_path, apk_filename + '.csv')
    csv_df = pd.read_csv(csv_path)

    # --- Load GAM and GAT attentions ---
    gam_df = get_and_process_attention_df(gam_attention_path, apk_filename, number_of_top_apis)
    gam_df.rename(columns={"attention": "GAM attention"}, inplace=True)
    gat_df = get_and_process_attention_df(gat_attention_path, apk_filename, number_of_top_apis)
    gat_df.rename(columns={"attention": "GAT attention"}, inplace=True)
    gam_gat_df = pd.merge(gam_df, gat_df, on='API Name', how='outer')

    # --- Merge APIs with methods ---
    merged_df = pd.merge(csv_df, gam_gat_df, left_on='opcode_output', right_on='API Name', how='inner')
    merged_df.drop('opcode_output', axis=1, inplace=True)
    merged_df = add_missing_methods(merged_df, custom_methods_path, apk_filename)
    merged_df = merged_df[~merged_df['method_name'].str.startswith(excluded_classes)]
    merged_df = merged_df.drop_duplicates()

    # === Method-level ===
    method_df = merged_df.groupby('method_name')[['GAM attention', 'GAT attention']].sum().reset_index()
    method_df = normalize_attentions(method_df)
    method_df = add_real_mcl_label(method_df, 'method_name', real_malicious_method_set)

    # === Class-level ===
    class_df = method_df.copy()
    class_df['class_name'] = class_df['method_name'].str.extract(r'(.+);->').apply(lambda x: x + "'")
    class_df = class_df[['class_name', 'normalized GAM attention', 'normalized GAT attention']]
    class_df = class_df.groupby('class_name')[
        ['normalized GAM attention', 'normalized GAT attention']].sum().reset_index()
    class_df = class_df.sort_values(by='normalized GAM attention', ascending=False)
    class_df = class_df[~class_df['class_name'].str.startswith(excluded_classes)]
    class_df = add_real_mcl_label(class_df, 'class_name', real_malicious_class_set)

    return gam_gat_df, method_df, class_df


def get_real_malicious_set(path, file):
    malicious_set = set()
    mystique_path = os.path.join(path, file + '.txt')
    with open(mystique_path, 'r', encoding='utf-8') as mystique_file:
        for line in mystique_file:
            malicious_set.add(line.strip())
    return malicious_set


def get_custom_method_set(path, file):
    method_set = set()
    method_path = os.path.join(path, file + '.txt')
    with open(method_path, 'r', encoding='utf-8') as file:
        for line in file:
            method_set.add(line.strip())
    return method_set


def get_and_process_attention_df(path, file, number):
    gam_excel_path = os.path.join(path, file + '.xlsx')
    dataframe = pd.read_excel(gam_excel_path, sheet_name='Sheet1')
    dataframe = dataframe.dropna(subset=["Number of Edges"])
    dataframe = dataframe[dataframe["Number of Edges"] != 0]
    dataframe.rename(columns={'Avg. Attention of Agents Classifying Mw': 'attention'}, inplace=True)
    dataframe = dataframe[['API Name', 'attention']]
    dataframe.sort_values(by='attention', ascending=False, inplace=True)
    dataframe.reset_index(drop=True, inplace=True)

    if isinstance(number, int):
        if number < len(dataframe.index):
            dataframe.loc[dataframe.index[number]:, 'attention'] = 0
    elif isinstance(number, float) and (number != 1.0):
        row_number = math.ceil(number * len(dataframe))
        dataframe.loc[dataframe.index[row_number]:, 'attention'] = 0

    return dataframe


def add_missing_methods(df, path, file):
    # Read the method names from the text file
    method_set = get_custom_method_set(path, file)

    # Find the missing method names
    missing_methods = method_set - set(df['method_name'])

    # Create a DataFrame with the missing methods and attention=0
    missing_df = pd.DataFrame({'method_name': list(missing_methods),
                               'GAM attention': [0] * len(missing_methods),
                               'GAT attention': [0] * len(missing_methods)})

    # Concatenate the original DataFrame and the DataFrame with missing methods
    updated_df = pd.concat([df, missing_df], ignore_index=True)

    return updated_df


def normalize_attentions(dataframe):
    # Create normalized columns
    dataframe['normalized GAM attention'] = dataframe['GAM attention'] / dataframe['GAM attention'].sum()
    dataframe['normalized GAT attention'] = dataframe['GAT attention'] / dataframe['GAT attention'].sum()

    # Drop original columns
    dataframe = dataframe.drop(['GAM attention', 'GAT attention'], axis=1)

    # Sort by 'normalized GAM attention'
    dataframe = dataframe.sort_values(by='normalized GAM attention', ascending=False)

    # Reset index
    dataframe = dataframe.reset_index(drop=True)

    return dataframe


def add_real_mcl_label(dataframe, mcl_type, real_malicious_set):
    dataframe[mcl_type] = dataframe[mcl_type].str.replace("'", '')
    dataframe['true_label'] = dataframe[mcl_type].str.strip().isin(item.strip() for item in real_malicious_set)
    dataframe['true_label'] = dataframe['true_label'].map({True: 'P', False: 'N'})

    return dataframe


def load_excel_data(calculation_dir, sheet_name):
    """
    Load all Excel files from a directory (including subfolders),
    extract the specified sheet, and concatenate them into one DataFrame.
    """
    # Search recursively for .xlsx files
    all_files = glob.glob(os.path.join(calculation_dir, "**", "*.xlsx"), recursive=True)

    df_list = []
    for f in all_files:
        df = pd.read_excel(f, sheet_name=sheet_name)
        df["apk_file"] = os.path.basename(f)  # track source file
        df_list.append(df)

    return pd.concat(df_list, ignore_index=True)


def compute_metrics(y_true, y_pred):
    """
    Compute standard classification metrics.
    """
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    metrics = {
        "Accuracy": accuracy_score(y_true, y_pred),
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "Recall": recall_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "FPR": fp / (fp + tn),
        "FNR": fn / (fn + tp),
    }
    return metrics


def plot_threshold(y_true, df, plot_filepath, txt_filepath, mcl_level, min_acceptable_recall_rate):
    thresholds = [0.00001, 0.00002, 0.00003, 0.00004, 0.00005, 0.00006, 0.00007, 0.00008, 0.00009,
                  0.0001, 0.0002, 0.0003, 0.0004, 0.0005, 0.0006, 0.0007, 0.0008, 0.0009,
                  0.001, 0.002, 0.003, 0.004, 0.005, 0.006, 0.007, 0.008, 0.009,
                  0.01]

    metrics_dict = {
        "Accuracy": [],
        "Precision": [],
        "Recall": [],
        "F1": [],
    }

    lines = []  # collect results for txt file

    for t in thresholds:
        preds = (df >= t).astype(int)
        acc = accuracy_score(y_true, preds)
        prec = precision_score(y_true, preds, zero_division=0)
        rec = recall_score(y_true, preds, zero_division=0)
        f1 = f1_score(y_true, preds, zero_division=0)

        metrics_dict["Accuracy"].append(acc)
        metrics_dict["Precision"].append(prec)
        metrics_dict["Recall"].append(rec)
        metrics_dict["F1"].append(f1)

        lines.append(f"Threshold={t:.5f} | Acc={acc:.4f}, Prec={prec:.4f}, Rec={rec:.4f}, F1={f1:.4f}")

    # Save results to txt file
    with open(txt_filepath, "w") as f:
        f.write("\n".join(lines))

    # Best threshold = max F1, but only if Recall >= min_acceptable_recall_rate
    valid_indices = [i for i, r in enumerate(metrics_dict["Recall"]) if r >= min_acceptable_recall_rate]

    if valid_indices:  # at least one threshold satisfies the recall constraint
        best_idx = max(valid_indices, key=lambda i: metrics_dict["F1"][i])
        best_threshold = thresholds[best_idx]
    else:
        # fallback: no threshold satisfies recall >= 0.98
        best_idx = np.argmax(metrics_dict["F1"])
        best_threshold = thresholds[best_idx]
        print("No threshold achieved recall >= 0.98. Falling back to best F1 overall.")

    # ---- Plot all metrics
    plt.figure(figsize=(8, 6))
    for metric, values in metrics_dict.items():
        plt.plot(thresholds, values, marker="o", label=metric)

    plt.axvline(
        best_threshold,
        color="red",
        linestyle="--",
        label=f"Selected threshold: {best_threshold:.5f}"
    )

    plt.xscale("log")
    plt.ylim(0.6, 1.0)  # y-axis between 60% and 100%
    plt.xlabel(f"{mcl_level.capitalize()} Level Attention Threshold (log scale)")
    plt.ylabel("Score")
    plt.legend()
    plt.tight_layout()
    plt.savefig(plot_filepath)
    plt.close()

    return best_threshold


def tune_weight_with_fusion(y_true, gam, gat, threshold, filepath, weights=np.linspace(0, 1, 50)):
    f1_scores = []
    for w in weights:
        fused = w * gam + (1 - w) * gat
        preds = (fused >= threshold).astype(int)
        f1_scores.append(f1_score(y_true, preds, zero_division=0))

    best_idx = np.argmax(f1_scores)
    best_weight = weights[best_idx]

    # Plot
    plt.figure(figsize=(7, 5))
    plt.plot(weights, f1_scores, marker="o", label="F1 Score (Fusion)")
    plt.axvline(best_weight, color="red", linestyle="--", label=f"Best={best_weight:.2f}")
    plt.xlabel("GAM Weight (w)")
    plt.ylabel("F1 Score")
    plt.title(f"Fusion F1 vs. GAM Weight (Threshold={threshold:.4f})")
    plt.legend()
    plt.savefig(filepath)
    plt.close()

    return best_weight


def plot_roc_pr_curves(y_true,
                       scores_dict,
                       gam_gat_localization_roc_curve_filepath,
                       gam_gat_localization_pr_curve_filepath,
                       extra_points=None):
    """
    Plot ROC and PR curves for multiple scoring strategies.
    scores_dict = { "Method name": predicted_scores_array }
    extra_points = { "Method name": binary_predictions_array }
    """
    # ---- ROC Curve
    plt.figure(figsize=(6, 5))
    for label, scores in scores_dict.items():
        fpr, tpr, _ = roc_curve(y_true, scores)
        roc_auc = auc(fpr, tpr)
        plt.plot(fpr, tpr, label=f"{label} (AUC={roc_auc:.3f})")

    # Extra single points (e.g., AND / OR rules)
    if extra_points is not None:
        for label, preds in extra_points.items():
            tn, fp, fn, tp = confusion_matrix(y_true, preds).ravel()
            fpr = fp / (fp + tn)
            tpr = tp / (tp + fn)
            plt.plot(fpr, tpr, "o", label=f"{label} (point)")

    plt.plot([0, 1], [0, 1], "k--")
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.legend()
    plt.savefig(gam_gat_localization_roc_curve_filepath)
    plt.close()

    # ---- PR Curve
    plt.figure(figsize=(6, 5))
    for label, scores in scores_dict.items():
        precision, recall, _ = precision_recall_curve(y_true, scores)
        pr_auc = average_precision_score(y_true, scores)
        plt.plot(recall, precision, label=f"{label} (AP={pr_auc:.3f})")

    if extra_points is not None:
        for label, preds in extra_points.items():
            p = precision_score(y_true, preds, zero_division=0)
            r = recall_score(y_true, preds, zero_division=0)
            plt.plot(r, p, "o", label=f"{label} (point)")

    plt.xlabel("Recall")
    plt.ylabel("Precision")
    plt.legend()
    plt.savefig(gam_gat_localization_pr_curve_filepath)
    plt.close()


def threshold_rule(scores, threshold):
    return (scores >= threshold).astype(int)


def fusion_rule(gam, gat, w):
    return w * gam + (1 - w) * gat


def evaluate_pipeline(mcl_level_column, df, gam, gat, y_true,
                      attention_threshold,
                      gam_weight,
                      gam_gat_localization_metrics_filepath,
                      gam_gat_localization_roc_curve_filepath,
                      gam_gat_localization_pr_curve_filepath,
                      gam_gat_localization_filepath
                      ):
    # Nice-to-have naming for output
    df["apk_name"] = df["apk_file"].str.replace(r"\.xlsx$", "", regex=True)
    df["GAM_normalized_attention"] = df["normalized GAM attention"]
    df["GAT_normalized_attention"] = df["normalized GAT attention"]

    fused = fusion_rule(gam, gat, w=gam_weight)

    results = {}

    # ---- Threshold-based predictions
    df["GAM_predicted_label"] = threshold_rule(gam, attention_threshold)
    df["GAT_predicted_label"] = threshold_rule(gat, attention_threshold)

    # AND fusion
    df["GAM_AND_GAT_predicted_label"] = (
        (df["GAM_predicted_label"] & df["GAT_predicted_label"])
    )

    # OR fusion
    df["GAM_OR_GAT_predicted_label"] = (
        (df["GAM_predicted_label"] | df["GAT_predicted_label"])
    )

    # Weighted fusion
    df["GAM_GAT_fusion_predicted_label"] = threshold_rule(fused, attention_threshold)

    # ---- Metrics (global over all rows)
    results["GAM"] = compute_metrics(y_true, df["GAM_predicted_label"])
    results["GAT"] = compute_metrics(y_true, df["GAT_predicted_label"])
    results["GAM AND GAT"] = compute_metrics(y_true, df["GAM_AND_GAT_predicted_label"])
    results["GAM OR GAT"] = compute_metrics(y_true, df["GAM_OR_GAT_predicted_label"])
    results["GAM-GAT Weighted Fusion"] = compute_metrics(y_true, df["GAM_GAT_fusion_predicted_label"])

    # ---- ROC & PR curves (use raw scores)
    scores_dict = {
        "GAM": gam,
        "GAT": gat,
        "Fusion": fused,
    }
    extra_points = {
        "GAM AND GAT": df["GAM_AND_GAT_predicted_label"],
        "GAM OR GAT": df["GAM_OR_GAT_predicted_label"],
    }
    plot_roc_pr_curves(y_true,
                       scores_dict,
                       gam_gat_localization_roc_curve_filepath,
                       gam_gat_localization_pr_curve_filepath,
                       extra_points=extra_points)

    # ---- Save metrics (txt)
    metrics_df = pd.DataFrame(results).T
    with open(gam_gat_localization_metrics_filepath, "w") as f:
        f.write(metrics_df.to_string())

    # ---- Save per-sample predictions (Excel)
    prediction_columns = [
        "apk_name",
        mcl_level_column,
        "GAM_normalized_attention",
        "GAT_normalized_attention",
        "true_label",
        "GAM_predicted_label",
        "GAT_predicted_label",
        "GAM_AND_GAT_predicted_label",
        "GAM_OR_GAT_predicted_label",
        "GAM_GAT_fusion_predicted_label",
    ]
    df[prediction_columns].to_excel(
        gam_gat_localization_filepath, index=False, sheet_name="predictions"
    )

    return metrics_df


def localize_malicious_codes(mcl_calculation_dir,
                             mcl_level,
                             mcl_metrics_vs_threshold_plot_filepath,
                             mcl_metrics_vs_threshold_txt_filepath,
                             localization_filepath,
                             localization_metrics_filepath,
                             localization_roc_curve_filepath,
                             localization_pr_curve_filepath,
                             min_acceptable_recall_rate
                             ):
    mcl_level_column = mcl_level + "_name"
    sheet_name = mcl_level + "_loc"

    # ---- Load data first
    df = load_excel_data(mcl_calculation_dir, sheet_name)
    df["true_label"] = df["true_label"].map({"P": 1, "N": 0})
    gam = df["normalized GAM attention"].values
    gat = df["normalized GAT attention"].values
    y_true = df["true_label"].values

    # ---- Step 1: Tune threshold using GAM F1
    attention_threshold = plot_threshold(y_true, gam, mcl_metrics_vs_threshold_plot_filepath,
                                         mcl_metrics_vs_threshold_txt_filepath, mcl_level, min_acceptable_recall_rate)
    print(f"Best GAM threshold for {mcl_level}: {attention_threshold:.5f}")

    # ---- Step 2: Final evaluation
    fusion_weight = 0.5
    evaluate_pipeline(mcl_level_column, df, gam, gat, y_true,
                      attention_threshold,
                      fusion_weight,
                      localization_metrics_filepath,
                      localization_roc_curve_filepath,
                      localization_pr_curve_filepath,
                      localization_filepath)
